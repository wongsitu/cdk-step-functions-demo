import boto3
import openpyxl
from io import BytesIO
from openpyxl import Workbook
import os
import uuid

s3 = boto3.client('s3')
bucket_name = os.environ['BUCKET_NAME']

def handler(event, context):
    results = event['results']
    
    # Create a new workbook for the merged file
    merged_workbook = Workbook()
    merged_workbook.remove(merged_workbook.active)  # Remove the default sheet
    
    # List to track S3 object keys for deletion later
    s3_keys_to_delete = []

    for result in results:
        url = result['presigned_url']
        s3_key = result['s3_key']
        # Extract bucket name and key from the S3 URL
        s3_keys_to_delete.append({'Key': s3_key})
        
        # Download the Excel file from S3
        response = s3.get_object(Bucket=bucket_name, Key=s3_key)
        excel_data = BytesIO(response['Body'].read())
        
        # Load the downloaded Excel file
        workbook = openpyxl.load_workbook(excel_data, data_only=True)
        
        # Copy each sheet from the downloaded workbook to the merged workbook
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            new_sheet = merged_workbook.create_sheet(title=sheet_name)
            
            # Copy data from the original sheet to the new sheet
            for row in sheet.iter_rows(values_only=True):
                new_sheet.append(row)

    # Save the merged workbook to a BytesIO object
    merged_file = BytesIO()
    merged_workbook.save(merged_file)
    merged_file.seek(0)
    
    # Define the key for the merged file
    file_uuid = uuid.uuid4()
    s3_key = f"{file_uuid}.xlsx"

    # Upload the merged file to S3
    s3.put_object(Bucket=bucket_name, Key=s3_key, Body=merged_file)
    
    # Delete the original files from S3
    if s3_keys_to_delete:
        s3.delete_objects(Bucket=bucket_name, Delete={'Objects': s3_keys_to_delete})
    
    # Generate a presigned URL for the merged file
    presigned_url = s3.generate_presigned_url(
        'get_object',
        Params={'Bucket': bucket_name, 'Key': s3_key},
        ExpiresIn=3600  # URL valid for 1 hour
    )
    
    return {
        'presignedUrl': presigned_url,
        'message': 'Merged file created and original files deleted successfully.'
    }
